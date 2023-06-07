import openpyxl
import requests
import pickle
import logging
from http.client import HTTPConnection
import re
import json
import time


class Donation:

    def __init__(self):
        self.amount = 0
        self.name = ""
        self.address = ""
        self.nationality = ""
        self.pin = 0
        self.country = ""
        self.state = ""
        self.city = ""
        self.pan = ""
        self.email = ""
        self.mobile = 0
        self.transaction_number = 0
        self.ip = requests.get('https://ipapi.co/ip/').text # Possibly expensive. But I dont think any more complexity is needed for this project at this time
        self._transaction_date = ""


    @property
    def transaction_date(self):
        return self._transaction_date
    
    @transaction_date.setter
    def transaction_date(self, transaction_date):
        if transaction_date != None:
            try:
                self._transaction_date = transaction_date.strftime("%Y-%m-%d")
            except:
                logging.error(f"Transaction date is not 'date' type. {transaction_date}. Please correct the transaction date in the excel and try again. Exiting Program.")
                exit()
        else:
            logging.error("Transaction date is None. Please correct transaction date and try again. Exiting Program")
            exit()

    def validate(self):
        if self.amount == None or self.amount < 500:
            logging.error("Donation anount should not be less then Rs. 500")
            return False
        elif self.name == None or len(self.name) < 1:
            logging.error("Donor name is None.")
            return False
        elif self.address == None or len(self.address) < 1:
            logging.error("Donor address name is None.")
            return False
        elif self.nationality == None or len(self.nationality) < 1:
            logging.error("Donor nationality is None.")
            return False
        elif self.pin == None or len(str(self.pin)) != 6:
            logging.error(f"Donor pin code is invalid. Pin: {self.pin}")
            return False
        elif self.country == None or len(self.country) < 1:
            logging.error(f"Donor country is invalid.")
            return False
        elif self.state == None or len(self.state) < 1:
            logging.error(f"Donor state is invalid.")
            return False
        elif self.city == None or len(self.city) < 1:
            logging.error(f"Donor city is invalid.")
            return False
        elif self.pan == None or len(self.pan) != 10:
            logging.error(f"Donor pan is invalid. Pan {self.pan}")
            return False
        elif self.email == None or len(self.email) < 1:
            logging.error(f"Donor email is invalid. Email: {self.email}")
            return False
        elif self.mobile == None or len(str(self.mobile)) != 10:
            logging.error(f"Donor mobile is invalid. Mobile: {self.mobile}")
            return False
        elif self.transaction_number == None:
            logging.error(f"Donation transacton number is invalid. Transaction #: {self.transaction_number}")
            return False
        return True
        
        

extensive_logging_enabled = False


def main():
    if extensive_logging_enabled:
        log = logging.getLogger('urllib3')
        log.setLevel(logging.DEBUG)

        # logging from urllib3 to console
        ch = logging.StreamHandler()
        ch.setLevel(logging.DEBUG)
        log.addHandler(ch)

        # print statements from `http.client.HTTPConnection` to console/stdout
        HTTPConnection.debuglevel = 1
    
    logging.basicConfig(format='%(asctime)s [%(levelname)s] - %(message)s', level=logging.INFO)


    # Read from Excel
    
    excel_filename = "donation.xlsx"
    logging.debug(f"Loading excel: {excel_filename}")
    wb_obj = openpyxl.load_workbook(excel_filename)
    logging.debug("Opening  active  excel sheet")
    sheet_obj = wb_obj.active
    row_count = sheet_obj.max_row
    logging.info(f"Row count(including headers): {row_count}")
    for i in range(2, row_count + 1):
        donation = Donation()
        rowSerialNumber = sheet_obj.cell(row = i, column = 1)
        if rowSerialNumber.value == None or rowSerialNumber.value == "":
            logging.info(f"Possible empty row found at excel row count: {i}. No serial number found. Assuming end of file and exiting.")
            exit()
        status = sheet_obj.cell(row = i, column = 16)
        donation.status = status.value
        if donation.status != "COMPLETED" and donation.status != "FAILED":
            amount = sheet_obj.cell(row = i, column = 2)
            donation.amount = amount.value
            name = sheet_obj.cell(row = i, column = 3)
            donation.name = name.value
            nationality = sheet_obj.cell(row = i, column = 4)
            donation.nationality = nationality.value
            address = sheet_obj.cell(row = i, column = 5)
            donation.address = address.value
            pin = sheet_obj.cell(row = i, column = 6)
            donation.pin = pin.value
            country = sheet_obj.cell(row = i, column = 7)
            donation.country = country.value
            state = sheet_obj.cell(row = i, column = 8)
            donation.state = state.value
            city = sheet_obj.cell(row = i, column = 9)
            donation.city = city.value
            pan = sheet_obj.cell(row = i, column = 10)
            donation.pan = pan.value
            email = sheet_obj.cell(row = i, column = 11)
            donation.email = email.value
            mobile = sheet_obj.cell(row = i, column = 12)
            donation.mobile = mobile.value
            transaction_number = sheet_obj.cell(row = i, column = 13)
            donation.transaction_number = transaction_number.value
            transaction_date = sheet_obj.cell(row = i, column = 14)
            donation.transaction_date = transaction_date.value
            reference_number = sheet_obj.cell(row = i, column = 15)
            donation.reference_number = reference_number.value


            # Validate excel row
            if not donation.validate():
                logging.error(f"Validation failed for row: {i}. Correct error above and try again. Exiting Program.")
                exit()

            openIframe()
            updateDonationAmount(amount.value)
            csrf_token, dm_token, csrf_payment_token = loadFormPage()
            donation.csrf_token = csrf_token
            donation.dm_token = dm_token
            donation.csrf_payment_token = csrf_payment_token
            if donation.csrf_token == None or donation.dm_token == None or donation.csrf_payment_token == None:
                sheet_obj.cell(row=i, column=16).value = "FAILED"
                logging.info(f"Row: {i} , donor: {donation.name} Failed")
            else:
                form_content = frameFormResponse(donation)
                donation_info_id = updateFormContents(form_content)
                updateUserAgent(donation_info_id)
                selectBankTransfer(donation_info_id)
                donation_reference_number = submitPaymentInfo(csrf_payment_token, donation_info_id, donation.transaction_date, donation.transaction_number)
                donation.donation_reference_number = donation_reference_number

                sheet_obj.cell(row=i, column=15).value = donation.donation_reference_number
                sheet_obj.cell(row=i, column=16).value = "COMPLETED"
                logging.info(f"Row: {i} , donor: {donation.name} updated successfully")
            wb_obj.save(excel_filename)
            
        else:
            logging.info(f"Not considering an already completed or failed row: {i}")

def openIframe():
    logging.debug("Opening danamojo iframe")
    headers = {
        'Content-Type': 'text/html; charset=utf-8',
        'Accept' : 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Referer' : 'https://www.gubbachi.org.in/',
        'Sec-Ch-Ua' : '"Google Chrome";v="113", "Chromium";v="113", "Not-A.Brand";v="24"',
        'Sec-Ch-Ua-Platform' : 'macOS',
        'Sec-Fetch-Dest' : 'iframe',
        'Sec-Fetch-Mode' : 'navigate',
        'Sec-Fetch-Site' : 'cross-site',
        'User-Agent' : 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36'
        }
    sendGet("https://www-gubbachi-org-in.filesusr.com/html/36d8e8_654919906176421ba7fb5a7e28a3301b.html", headers, False)
    

def updateDonationAmount(donation_amount):
    
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'Host' : 'danamojo.org',
        'Origin' : 'https://www-gubbachi-org-in.filesusr.com',
        'Host' : 'danamojo.org',
        'Referer' : 'https://www.gubbachi.org.in/',
        'Sec-Ch-Ua' : '"Google Chrome";v="113", "Chromium";v="113", "Not-A.Brand";v="24"',
        'Sec-Fetch-Mode' : 'cors',
        'Sec-Fetch-Site' : 'navigate',
        'Sec-Fetch-Site' : 'cross-site',
        'User-Agent' : 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36'
    }

    data = f"refererUrl=https%3A%2F%2Fwww.gubbachi.org.in%2F&totalPrice={donation_amount}&ngoId=959"
    logging.debug(f"Updating donation amount. Request: {data}")
    sendPost('https://danamojo.org/dm/widget/update-logs', headers, data)


def loadFormPage():

    headers = {
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'Host' : 'danamojo.org',
        'Origin' : 'https://www-gubbachi-org-in.filesusr.com',
        'Host' : 'danamojo.org',
        'Referer' : 'https://www.gubbachi.org.in/',
        'Sec-Ch-Ua' : '"Google Chrome";v="113", "Chromium";v="113", "Not-A.Brand";v="24"',
        'Sec-Fetch-Mode' : 'cors',
        'Sec-Fetch-Site' : 'navigate',
        'Sec-Fetch-Site' : 'cross-site',
        'User-Agent' : 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36'
    }
    logging.debug(f"Loading form page")
    content , status = sendGet("https://danamojo.org/dm/widget/index?ngoId=959&products=&layoutType=GRID&currencyCode=&ip_country=&device=Desktop&oDisplayTab=once,monthly&sw=1792&sourceType=3&refererUrl=https%3A%2F%2Fwww.gubbachi.org.in%2F&utm_campaign=&qrCode=YES", headers, True)
    #print("Printing srf content")
    #print(content)
    csrf_token = captureCsrfToken(content)
    logging.debug(f"csrf_token : {csrf_token}")
    dm_token = captureDmToken(content)
    logging.debug(f"_dn_token : {dm_token}")
    csrf = captureCsrfForPayment(content)
    logging.debug(f"csrf payment token : {csrf}")
    return csrf_token, dm_token, csrf
    
def captureCsrfToken(content):
    logging.debug(f"Capturing CSRF token")
    regex = "csrf-token.*>"
    match = re.search(regex, content)
    if match:
        sub_string = match.group()
        csrf_token = sub_string.split('"')[2]
        return csrf_token
    else:
        logging.error(f"CSRF Token not found. Exiting now")
        return None

def captureDmToken(content):
    logging.debug(f"Capturing _dmToken")
    regex = "_dmToken.*>(.*)<\/span>"
    match = re.search(regex, content)
    if match:
        sub_string = match.group(1)
        dm_token = sub_string
        return dm_token
    else:
        print("_dmToken Token not found. Exiting now")
        exit

def captureCsrfForPayment(content):
    logging.debug(f"Capturing Csrf For Payment")
    #regex = "name=\"_csrf\"( value=\"value=\")(.*)\">"
    regex = "_csrf"
    match = re.search(regex, content)
    if match:
        sub_string = match.group()
        csrf = sub_string
        return csrf
    else:
        logging.error(f"_csrf Token not found. Exiting now")
        return None

def frameFormResponse(donation):
    
    form_response = ( f"_csrf={donation.csrf_token}"
                     "&product_qty=0"
                     "&productInfo%5B7164%5D%5BdonationProductType%5D=2"
                     "&productInfo%5B7164%5D%5BdonationProductName%5D=Train+a+teacher+for+our+specialised+education+programmes"
                     "&productInfo%5B7164%5D%5BdonationProductPrice%5D=16500"
                     "&productInfo%5B7164%5D%5BdonationProductAmount%5D="
                     "&productInfo%5B7164%5D%5BdonationProductQty%5D=0"
                     "&productInfo%5B7164%5D%5BdonationProductTaxBenefit%5D=50"
                     "&productInfo%5B7164%5D%5BdonationProductNGOId%5D=959"
                     "&product_dm_price=0"
                     "&productInfo%5B7166%5D%5BdonationProductType%5D=3"
                     "&productInfo%5B7166%5D%5BdonationProductName%5D=Contribute+towards+our+educational+initiatives"
                     "&productInfo%5B7166%5D%5BdonationProductPrice%5D=2000"
                     "&productInfo%5B7166%5D%5BdonationProductAmount%5D=0"
                     "&productInfo%5B7166%5D%5BdonationProductQty%5D=0"
                     "&productInfo%5B7166%5D%5BdonationProductTaxBenefit%5D=50"
                     "&productInfo%5B7166%5D%5BdonationProductNGOId%5D=959"
                     "&product_qty=0"
                     "&productInfo%5B5784%5D%5BdonationProductType%5D=2"
                     "&productInfo%5B5784%5D%5BdonationProductName%5D=Sponsor+a+child's+education+for+one+month"
                     "&productInfo%5B5784%5D%5BdonationProductPrice%5D=1750"
                     "&productInfo%5B5784%5D%5BdonationProductAmount%5D="
                     "&productInfo%5B5784%5D%5BdonationProductQty%5D=0"
                     "&productInfo%5B5784%5D%5BdonationProductTaxBenefit%5D=50"
                     "&productInfo%5B5784%5D%5BdonationProductNGOId%5D=959"
                     "&product_dm_price=0"
                     "&productInfo%5B7167%5D%5BdonationProductType%5D=3"
                     "&productInfo%5B7167%5D%5BdonationProductName%5D=Contribute+towards+our+nutritional+initiatives"
                     "&productInfo%5B7167%5D%5BdonationProductPrice%5D=1000"
                     "&productInfo%5B7167%5D%5BdonationProductAmount%5D=0"
                     "&productInfo%5B7167%5D%5BdonationProductQty%5D=0"
                     "&productInfo%5B7167%5D%5BdonationProductTaxBenefit%5D=50"
                     "&productInfo%5B7167%5D%5BdonationProductNGOId%5D=959"
                     "&product_qty=0"
                     "&productInfo%5B7163%5D%5BdonationProductType%5D=2"
                     "&productInfo%5B7163%5D%5BdonationProductName%5D=Help+our+community+workers+and+educators+stay+dry+in+wet+weather"
                     "&productInfo%5B7163%5D%5BdonationProductPrice%5D=800"
                     "&productInfo%5B7163%5D%5BdonationProductAmount%5D="
                     "&productInfo%5B7163%5D%5BdonationProductQty%5D=0"
                     "&productInfo%5B7163%5D%5BdonationProductTaxBenefit%5D=50"
                     "&productInfo%5B7163%5D%5BdonationProductNGOId%5D=959"
                     f"&product_dm_price={donation.amount}"
                     "&productInfo%5B5720%5D%5BdonationProductType%5D=3"
                     "&productInfo%5B5720%5D%5BdonationProductName%5D=CSR+Activities"
                     f"&productInfo%5B5720%5D%5BdonationProductPrice%5D={donation.amount}"
                     f"&productInfo%5B5720%5D%5BdonationProductAmount%5D={donation.amount}"
                     "&productInfo%5B5720%5D%5BdonationProductQty%5D=0"
                     "&productInfo%5B5720%5D%5BdonationProductTaxBenefit%5D=50"
                     "&productInfo%5B5720%5D%5BdonationProductNGOId%5D=959"
                     "&has80GForTxnCharges=1"
                     "&donorTxnCharges=0"
                     f"&DonationInfo[fullName]={donation.name}"
                     f"&DonationInfo[nationality]={donation.nationality}"
                     f"&DonationInfo[address]={donation.address}"
                     f"&DonationInfo[pincode]={donation.pin}"
                     f"&DonationInfo[country]={donation.country}"
                     f"&DonationInfo[state]={donation.state}"
                     f"&DonationInfo[city]={donation.city}"
                     f"&DonationInfo[id]={donation.pan}"
                     f"&DonationInfo[idProof]=1"
                     f"&DonationInfo[email]={donation.email}"
                     f"&DonationInfo[mobile]={donation.mobile}"
                     f"&countryCode=+91"
                     f"&DonationInfo[country]={donation.country}"
                     f"&DonationInfo[ipAddress]={donation.ip}"
                     f"&DonationInfo[passportNumber]="
                     f"&DonationInfo[passportCopyLink]="
                     f"&DonationInfo[challenge]="
                     "&customPassportfile="
                     "&passportFileAlreadyUploaded="
                     "&recurringDonation=0"
                     "&donorInfoId="
                     "&showID=1"
                     "&panVerification=0"
                     "&DonationInfo[sourceType]=3"
                     "&DonationInfo[iframeUrl]=https%3A%2F%2Fwww.gubbachi.org.in%2F"
                     f"&DonationInfo[ipAddress]={donation.ip}"
                     f"&product_dm_price={donation.amount}"
                     "&customDonationInfoId="
                     "&gifting_checkbox=0"
                     "&gifting_occasion="
                     "&gifting_occasions_gift_type=1"
                     "&gifting_recipient_name="
                     "&gifting_recipient_email="
                     "&gifting_subject_line="
                     "&gifting_message="
                     "&gifting_show_gifting_amount=0"
                     "&gifting_card_image="
                     "&gifting_upload_image="
                     "&DonationInfo[layoutType]="
                     "&DonationInfo[device]=Desktop"
                     f"&_dmToken={donation.dm_token}"
    )

    logging.debug(f"Form content: {form_response}")
    return form_response


def updateFormContents(form_content):
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'Host' : 'danamojo.org',
        'Origin' : 'https://www-gubbachi-org-in.filesusr.com',
        'Host' : 'danamojo.org',
        'Referer' : 'https://www.gubbachi.org.in/',
        'Sec-Ch-Ua' : '"Google Chrome";v="113", "Chromium";v="113", "Not-A.Brand";v="24"',
        'Sec-Fetch-Mode' : 'cors',
        'Sec-Fetch-Site' : 'navigate',
        'Sec-Fetch-Site' : 'cross-site',
        'User-Agent' : 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36'
    }

    data = form_content
    logging.debug(f"Update form contents. Request : {data}")
    content, status = sendPost("https://danamojo.org/dm/widget/create", headers, data)
    donaton_info_id = ""
    try:
        donaton_info_id = json.loads(content)["donationInfoId"]
    except (json.decoder.JSONDecodeError, KeyError):
        logging.error(f"Danamojo did not accept the donor data. Response : {content}, status: {status}")
        exit()
    logging.debug(f"Donation Id : {donaton_info_id}")
    return donaton_info_id

    

def updateUserAgent(donation_info_id):
   
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'Host' : 'danamojo.org',
        'Origin' : 'https://www-gubbachi-org-in.filesusr.com',
        'Host' : 'danamojo.org',
        'Referer' : 'https://www.gubbachi.org.in/',
        'Sec-Ch-Ua' : '"Google Chrome";v="113", "Chromium";v="113", "Not-A.Brand";v="24"',
        'Sec-Fetch-Mode' : 'cors',
        'Sec-Fetch-Site' : 'navigate',
        'Sec-Fetch-Site' : 'cross-site',
        'User-Agent' : 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36'
    }
    
    data = f"donationInfoId={donation_info_id}&userAgent=Browser+name++%3D+Chrome%2C+Full+version++%3D+113.0.0.0%2C+Major+version+%3D+113%2C+navigator.appName+%3D+Netscape%2C+navigator.userAgent+%3D+Mozilla%2F5.0+(Macintosh%3B+Intel+Mac+OS+X+10_15_7)+AppleWebKit%2F537.36+(KHTML%2C+like+Gecko)+Chrome%2F113.0.0.0+Safari%2F537.36"
    logging.debug(f"Update User Agent. Request : {data}")
    sendPost("https://danamojo.org/dm/widget/update-user-agent", headers, data)
    

def selectBankTransfer(donation_info_id):
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'Host' : 'danamojo.org',
        'Origin' : 'https://www-gubbachi-org-in.filesusr.com',
        'Host' : 'danamojo.org',
        'Referer' : 'https://www.gubbachi.org.in/',
        'Sec-Ch-Ua' : '"Google Chrome";v="113", "Chromium";v="113", "Not-A.Brand";v="24"',
        'Sec-Fetch-Mode' : 'cors',
        'Sec-Fetch-Site' : 'navigate',
        'Sec-Fetch-Site' : 'cross-site',
        'User-Agent' : 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36'
    }
    time_in_millis = time.time()
    logging.debug(f"Select bank transfer")
    sendGet(f"https://danamojo.org/dm/widget/loadpaymentform/?formName=bank_transfer&donationInfoId={donation_info_id}&payment_option=6&_={time_in_millis}", headers, True)
    

def submitPaymentInfo(csrf_token, donatino_info_id, donation_date, transaction_number ):

    headers = {
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'Host' : 'danamojo.org',
        'Origin' : 'https://www-gubbachi-org-in.filesusr.com',
        'Host' : 'danamojo.org',
        'Referer' : 'https://www.gubbachi.org.in/',
        'Sec-Ch-Ua' : '"Google Chrome";v="113", "Chromium";v="113", "Not-A.Brand";v="24"',
        'Sec-Fetch-Mode' : 'cors',
        'Sec-Fetch-Site' : 'navigate',
        'Sec-Fetch-Site' : 'cross-site',
        'User-Agent' : 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36'
    }


    data = ("_csrf=4CRcz_1bpyX6Im_WrEPwAW9lVcWthcEfT7gdUnitN-CmZTaBmR_tZ6MRVpOeFoVWNz0DkPzyqEc_zXYzKd5ntA%3D%3D"
            f"&DonationInfo[donationInfoId]={donatino_info_id}"
            f"&DonationInfo[chequeDraftDate]={donation_date}"
            f"&DonationInfo[trackingNumber]={transaction_number}"
            "&DonationInfo[formName]=bank_transfer"
            "&DonationInfo[paymentStatus]=1"
            "&DonationInfo[paymentOption]=6"
    )

    logging.debug(f"Submit payment. Data: {data}")
    content, status = sendPost("https://danamojo.org/dm/widget/updatepaymentdetails", headers, data)
    donation_refernce_string = json.loads(content)["data"]
    donation_reference_number = captureDonationReference(donation_refernce_string)
    logging.debug(f"The donation reference number: {donation_reference_number}")
    return donation_reference_number
    
def captureDonationReference(content):
    regex = "<strong>(.*)<\/strong>"
    match = re.search(regex, content)
    if match:
        sub_string = match.group(1)

        donation_reference_number = sub_string
        return donation_reference_number
    else:
        logging.error("Donation_reference_number not found. Exiting now")
        return None

def noThanks():
    logging.debug("Send Thanks message")


def sendPost(url, headers, data):
    resp = requests.post(url, cookies=loadCookies("cookies_py"), headers=headers, data=data)
    saveCookies(resp.cookies)
    logging.debug(f"Sending POST :{url} ")
    logging.debug(f"Response status: {resp.status_code} , Response content: {resp.text} ")
    return resp.text, resp.status_code

def sendGet(url, headers, loadCookiesFromFile):
    logging.debug(f"Sending GET :{url} ")
    if loadCookiesFromFile:
        resp = requests.get(url, cookies=loadCookies("cookies_py"), headers=headers)
        logging.debug(f"Saving Cookies")
        saveCookies(resp.cookies)
        logging.debug(f"Response status: {resp.status_code} , Response content: {resp.text} ")
        return resp.text, resp.status_code

    else:
        resp = requests.get(url, headers=headers)
        logging.debug(f"Saving Cookies")
        saveCookies(resp.cookies)
        logging.debug(f"Response status: {resp.status_code} , Response content: {resp.text} ")
        return resp.text, resp.status_code

def saveCookies(request_cookies):
    filename = "./cookies_py"
    with open(filename, 'wb') as f:
        pickle.dump(request_cookies, f)

def loadCookies(filename):
    with open(filename, 'rb') as f:
        return pickle.load(f)

if __name__ == "__main__":
    #logging.info("Starting Gubbachi automation")
    main()
